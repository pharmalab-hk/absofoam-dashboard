import io
from typing import Optional

import pandas as pd
import plotly.express as px
import requests
import streamlit as st
from openpyxl import load_workbook

# =========================================================
# Page config
# =========================================================
st.set_page_config(
    page_title="ABSOFOAM Dashboard",
    page_icon="📊",
    layout="wide"
)

# =========================================================
# Constants
# =========================================================
DEFAULT_LOCAL_FILE = r"Z:\PRODUCTS\Product Monitoring Trend\FORYOU\ABSOFOAM Adhesiveness Trend.xlsx"
DEFAULT_SHEET_NAME = "Data"

# Optional default URL from Streamlit secrets
DEFAULT_ONEDRIVE_URL = ""
if "onedrive" in st.secrets and "direct_url" in st.secrets["onedrive"]:
    DEFAULT_ONEDRIVE_URL = st.secrets["onedrive"]["direct_url"]

# =========================================================
# UI helpers
# =========================================================
def format_number(value: Optional[float], decimals: int = 2) -> str:
    if pd.isna(value):
        return "-"
    return f"{value:.{decimals}f}"


def format_percent(value: Optional[float], decimals: int = 1) -> str:
    if pd.isna(value):
        return "-"
    return f"{value:.{decimals}%}"


def info_box(message: str) -> None:
    st.info(message)


# =========================================================
# Data normalization
# =========================================================
def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Remove fully empty columns
    df = df.dropna(axis=1, how="all")

    # Normalize column names
    df.columns = [str(col).strip() for col in df.columns]

    # Standardize reference code naming if needed
    if "Reference Code" in df.columns and "Reference code" not in df.columns:
        df = df.rename(columns={"Reference Code": "Reference code"})

    # Numeric columns
    numeric_cols = [
        "Adhesiveness reading 1",
        "Adhesiveness reading 2",
        "Adhesiveness reading 3",
        "Adhesiveness on Inspection Report",
        "Adhesiveness on COA",
        "Discrepancy",
        "LOT#",
        "Year",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Text columns
    text_cols = [
        "Lot Number",
        "Product Range",
        "Reference code",
        "Remarks",
    ]
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # Clean fake "nan" strings created by astype(str)
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})

    # Remove rows that are empty on key identifiers
    key_cols = [col for col in ["Product Range", "Lot Number"] if col in df.columns]
    if key_cols:
        df = df.dropna(subset=key_cols, how="all")

    # Keep Year nullable integer if possible
    if "Year" in df.columns:
        df["Year"] = df["Year"].astype("Int64")

    return df


# =========================================================
# File loading
# =========================================================
def load_excel_data_from_bytes(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.values
    columns = next(rows)
    df = pd.DataFrame(rows, columns=columns)
    return normalize_dataframe(df)


def load_excel_data_from_path(path: str, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.values
    columns = next(rows)
    df = pd.DataFrame(rows, columns=columns)
    return normalize_dataframe(df)


def build_download_hint(url: str) -> str:
    if "download=1" in url:
        return url
    if "?" in url:
        return f"{url}&download=1"
    return f"{url}?download=1"


@st.cache_data(show_spinner=False)
def load_data_from_url(url: str, sheet_name: str) -> pd.DataFrame:
    """
    Attempts to load an Excel workbook from URL.
    Raises a clear error if the URL returns HTML instead of an XLSX file.
    """
    response = requests.get(
        url,
        timeout=45,
        allow_redirects=True,
        headers={"User-Agent": "Mozilla/5.0"}
    )
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()

    # Common sign that a sharing page / preview page is being returned
    if "text/html" in content_type:
        suggested_url = build_download_hint(url)
        raise ValueError(
            "The provided link returns a web page instead of the Excel file.\n\n"
            "Please use a direct-download link. You can try this first:\n"
            f"{suggested_url}"
        )

    # XLSX is a zip-based file format, usually starts with PK
    if not response.content[:2] == b"PK":
        raise ValueError(
            "The downloaded content is not a valid Excel .xlsx file.\n\n"
            "This usually means the OneDrive link is a preview/share page, not a direct file download."
        )

    return load_excel_data_from_bytes(response.content, sheet_name)


@st.cache_data(show_spinner=False)
def load_data_from_local(path: str, sheet_name: str) -> pd.DataFrame:
    return load_excel_data_from_path(path, sheet_name)


# =========================================================
# Validation
# =========================================================
def validate_required_columns(df: pd.DataFrame) -> None:
    required_columns = [
        "Year",
        "Lot Number",
        "Product Range",
        "Adhesiveness on Inspection Report",
        "Adhesiveness on COA",
    ]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        st.error(
            f"Colonnes manquantes / Missing required columns: {missing_columns}"
        )
        st.stop()


# =========================================================
# Header
# =========================================================
st.title("📊 ABSOFOAM – Adhesiveness Dashboard")
st.caption("Suivi interactif de l’adhésivité | Interactive adhesiveness monitoring")

# =========================================================
# Sidebar - Data source
# =========================================================
st.sidebar.header("Source des données / Data source")

data_source = st.sidebar.radio(
    "Choisir la source / Choose source",
    options=["OneDrive link", "Upload file", "Local Excel file"],
    index=0
)

sheet_name = st.sidebar.text_input(
    "Nom de feuille / Sheet name",
    value=DEFAULT_SHEET_NAME
)

df = None
load_error = None

try:
    if data_source == "OneDrive link":
        onedrive_url = st.sidebar.text_input(
            "Lien OneDrive / OneDrive URL",
            value=DEFAULT_ONEDRIVE_URL,
            help="Use a direct Excel file URL if possible. A normal share page may fail."
        )

        if onedrive_url:
            df = load_data_from_url(onedrive_url, sheet_name)
        else:
            info_box(
                "Collez un lien OneDrive direct, ou utilisez Upload file.\n\n"
                "Paste a direct OneDrive link, or use Upload file."
            )

    elif data_source == "Upload file":
        uploaded_file = st.sidebar.file_uploader(
            "Téléverser le fichier Excel / Upload Excel file",
            type=["xlsx"]
        )
        if uploaded_file is not None:
            df = load_excel_data_from_bytes(uploaded_file.read(), sheet_name)
        else:
            info_box(
                "Veuillez téléverser un fichier Excel pour continuer.\n\n"
                "Please upload an Excel file to continue."
            )

    elif data_source == "Local Excel file":
        local_path = st.sidebar.text_input(
            "Chemin local / Local path",
            value=DEFAULT_LOCAL_FILE
        )
        df = load_data_from_local(local_path, sheet_name)

except Exception as exc:
    load_error = str(exc)

# =========================================================
# Graceful load fallback messaging
# =========================================================
if load_error:
    st.error(f"Erreur de chargement / Loading error:\n\n{load_error}")

    if data_source == "OneDrive link":
        st.warning(
            "Suggestion:\n"
            "1. Try appending `&download=1` to your OneDrive link.\n"
            "2. If it still fails, switch to `Upload file` for now.\n"
            "3. Once upload works, we can refine the OneDrive direct-download setup."
        )

if df is None:
    st.stop()

# =========================================================
# Validate data
# =========================================================
validate_required_columns(df)

# =========================================================
# Sidebar - Filters
# =========================================================
st.sidebar.header("Filtres / Filters")

product_ranges = sorted(
    df["Product Range"].dropna().astype(str).unique().tolist()
)

years = sorted(
    df["Year"].dropna().astype(int).unique().tolist()
) if "Year" in df.columns else []

selected_product_ranges = st.sidebar.multiselect(
    "Gamme produit / Product Range",
    options=product_ranges,
    default=product_ranges[:1] if product_ranges else []
)

reference_df = df.copy()
if selected_product_ranges:
    reference_df = reference_df[
        reference_df["Product Range"].astype(str).isin(selected_product_ranges)
    ]

reference_codes = []
if "Reference code" in reference_df.columns:
    reference_codes = sorted(
        reference_df["Reference code"].dropna().astype(str).unique().tolist()
    )

selected_reference_codes = st.sidebar.multiselect(
    "Référence produit / Product Reference",
    options=reference_codes,
    default=reference_codes
)

selected_years = st.sidebar.multiselect(
    "Année / Year",
    options=years,
    default=years
)

metric_choice = st.sidebar.selectbox(
    "Métrique du graphique / Chart metric",
    options=["Inspection only", "COA only", "Both"],
    index=2
)

show_raw_data = st.sidebar.checkbox(
    "Afficher les données brutes / Show raw data",
    value=False
)

# =========================================================
# Filter data
# =========================================================
filtered_df = df.copy()

if selected_product_ranges:
    filtered_df = filtered_df[
        filtered_df["Product Range"].astype(str).isin(selected_product_ranges)
    ]

if selected_reference_codes and "Reference code" in filtered_df.columns:
    filtered_df = filtered_df[
        filtered_df["Reference code"].astype(str).isin(selected_reference_codes)
    ]

if selected_years and "Year" in filtered_df.columns:
    filtered_df = filtered_df[
        filtered_df["Year"].isin(selected_years)
    ]

if filtered_df.empty:
    st.warning(
        "Aucune donnée pour les filtres sélectionnés / No data for the selected filters."
    )
    st.stop()

# =========================================================
# Derived summaries
# =========================================================
total_rows = len(filtered_df)
total_lots = filtered_df["Lot Number"].nunique()
avg_inspection = filtered_df["Adhesiveness on Inspection Report"].mean()
avg_coa = filtered_df["Adhesiveness on COA"].mean()
avg_discrepancy = (
    filtered_df["Discrepancy"].mean()
    if "Discrepancy" in filtered_df.columns
    else pd.NA
)

chart_df = (
    filtered_df.groupby(["Year", "Lot Number"], as_index=False)[
        ["Adhesiveness on Inspection Report", "Adhesiveness on COA"]
    ]
    .mean()
    .sort_values(by=["Year", "Lot Number"])
)

chart_df["Lot Label"] = (
    chart_df["Year"].astype(str) + " | " + chart_df["Lot Number"].astype(str)
)

yearly_avg = (
    filtered_df.groupby("Year", as_index=False)[
        ["Adhesiveness on Inspection Report", "Adhesiveness on COA"]
    ]
    .mean()
)

yearly_long = yearly_avg.melt(
    id_vars="Year",
    value_vars=["Adhesiveness on Inspection Report", "Adhesiveness on COA"],
    var_name="Metric",
    value_name="Average Adhesiveness"
)

yearly_long["Metric"] = yearly_long["Metric"].replace({
    "Adhesiveness on Inspection Report": "Inspection Report",
    "Adhesiveness on COA": "COA"
})

# =========================================================
# Top summary row
# =========================================================
top_left, top_right = st.columns([4, 1])

with top_left:
    selected_products_text = ", ".join(selected_product_ranges) if selected_product_ranges else "All"
    st.markdown(
        f"**Gamme sélectionnée / Selected range:** {selected_products_text}"
    )

with top_right:
    if st.button("🔄 Actualiser / Refresh"):
        st.cache_data.clear()
        st.rerun()

# =========================================================
# KPI row
# =========================================================
k1, k2, k3, k4 = st.columns(4)
k1.metric("Inspection moyenne / Avg Inspection", format_number(avg_inspection, 2))
k2.metric("COA moyenne / Avg COA", format_number(avg_coa, 2))
k3.metric("Lots uniques / Unique Lots", f"{total_lots}")
k4.metric("Écart moyen / Avg Discrepancy", format_percent(avg_discrepancy, 1))

# =========================================================
# Tabs
# =========================================================
tab1, tab2, tab3 = st.tabs([
    "📈 Vue d'ensemble / Overview",
    "📊 Analyse / Analysis",
    "🧾 Données / Data"
])

# =========================================================
# TAB 1 - Overview
# =========================================================
with tab1:
    st.subheader("Tendance par lot / Trend by lot")

    if metric_choice == "Both":
        trend_long = chart_df.melt(
            id_vars=["Year", "Lot Number", "Lot Label"],
            value_vars=["Adhesiveness on Inspection Report", "Adhesiveness on COA"],
            var_name="Metric",
            value_name="Adhesiveness"
        )

        trend_long["Metric"] = trend_long["Metric"].replace({
            "Adhesiveness on Inspection Report": "Inspection Report",
            "Adhesiveness on COA": "COA"
        })

        fig_trend = px.line(
            trend_long,
            x="Lot Label",
            y="Adhesiveness",
            color="Metric",
            markers=True,
            hover_data=["Year", "Lot Number"],
            title="Inspection Report vs COA"
        )
    else:
        y_col = (
            "Adhesiveness on Inspection Report"
            if metric_choice == "Inspection only"
            else "Adhesiveness on COA"
        )

        chart_title = (
            "Inspection Report trend"
            if metric_choice == "Inspection only"
            else "COA trend"
        )

        fig_trend = px.line(
            chart_df,
            x="Lot Label",
            y=y_col,
            markers=True,
            hover_data=["Year", "Lot Number"],
            title=chart_title
        )

    fig_trend.update_layout(
        xaxis_title="Lot",
        yaxis_title="Adhesiveness",
        height=500
    )
    st.plotly_chart(fig_trend, use_container_width=True)

    c1, c2 = st.columns(2)

    with c1:
        st.subheader("Moyenne annuelle / Yearly average")
        fig_year = px.bar(
            yearly_long,
            x="Year",
            y="Average Adhesiveness",
            color="Metric",
            barmode="group",
            title="Average adhesiveness by year"
        )
        fig_year.update_layout(height=420)
        st.plotly_chart(fig_year, use_container_width=True)

    with c2:
        st.subheader("Écart annuel / Yearly discrepancy")
        if "Discrepancy" in filtered_df.columns:
            discrepancy_by_year = (
                filtered_df.groupby("Year", as_index=False)["Discrepancy"]
                .mean()
                .sort_values("Year")
            )

            fig_disc = px.bar(
                discrepancy_by_year,
                x="Year",
                y="Discrepancy",
                title="Average discrepancy by year"
            )
            fig_disc.update_layout(height=420)
            st.plotly_chart(fig_disc, use_container_width=True)
        else:
            st.info("La colonne 'Discrepancy' n’est pas disponible / 'Discrepancy' column not available.")

# =========================================================
# TAB 2 - Analysis
# =========================================================
with tab2:
    st.subheader("Résumé analytique / Analytical summary")

    a1, a2 = st.columns(2)

    with a1:
        st.markdown("**Périmètre / Scope**")
        st.write(f"- Lignes filtrées / Filtered rows: **{total_rows}**")
        st.write(f"- Lots uniques / Unique lots: **{total_lots}**")
        st.write(f"- Années sélectionnées / Selected years: **{', '.join(map(str, selected_years)) if selected_years else 'All'}**")

    with a2:
        st.markdown("**Indicateurs / Indicators**")
        st.write(f"- Inspection moyenne / Avg Inspection: **{format_number(avg_inspection, 2)}**")
        st.write(f"- COA moyenne / Avg COA: **{format_number(avg_coa, 2)}**")
        st.write(f"- Écart moyen / Avg Discrepancy: **{format_percent(avg_discrepancy, 1)}**")

    if "Reference code" in filtered_df.columns:
        st.subheader("Répartition par référence / Breakdown by reference")

        ref_summary = (
            filtered_df.groupby("Reference code", as_index=False)
            .agg(
                Lots=("Lot Number", "nunique"),
                Avg_Inspection=("Adhesiveness on Inspection Report", "mean"),
                Avg_COA=("Adhesiveness on COA", "mean")
            )
            .sort_values("Lots", ascending=False)
        )

        fig_ref = px.bar(
            ref_summary,
            x="Reference code",
            y="Lots",
            title="Number of lots by reference"
        )
        fig_ref.update_layout(height=420)
        st.plotly_chart(fig_ref, use_container_width=True)

        st.dataframe(ref_summary, use_container_width=True)

# =========================================================
# TAB 3 - Data
# =========================================================
with tab3:
    st.subheader("Données filtrées / Filtered data")

    if show_raw_data:
        st.dataframe(filtered_df, use_container_width=True)
    else:
        st.info(
            "Cochez l’option dans la barre latérale pour afficher les données brutes.\n\n"
            "Use the sidebar checkbox to display raw data."
        )

    csv_data = filtered_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Télécharger les données filtrées / Download filtered CSV",
        data=csv_data,
        file_name="absofoam_filtered_data.csv",
        mime="text/csv"
    )

# =========================================================
# Footer note
# =========================================================
st.markdown("---")
st.caption(
    "Tip: if OneDrive loading fails, use Upload file first. "
    "Once confirmed working, replace the share link with a direct-download link."
)
